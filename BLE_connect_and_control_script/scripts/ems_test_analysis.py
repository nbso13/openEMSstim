from numpy.core.fromnumeric import size
import ems_constants
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
from ems_test import spike_times_to_traces
from ems_test import process_contact_trace_to_hit_times
from ems_test import plot_contact_trace_and_rhythm
import glob
import quantities as pq
from elephant.spike_train_dissimilarity import victor_purpura_distance
from neo import SpikeTrain
import scipy


def plot_traces(x_array, trace_list, samp_period, legend_labels, title):
    fig, ax = plt.subplots()
    
    ax.set_yticks(np.arange(0, 500, 100))
    ax.set_xticks(np.arange(0, (len(trace_list[0]) * samp_period), 10000))
    ax.plot(x_array, trace_list[0])
    for i in range(len(trace_list) - 1):
        ax.plot(x_array, trace_list[i+1]*np.max(trace_list[0])/2)
    ax.legend(legend_labels)
    ax.set_title(title)
    plt.ion()
    plt.show()
    plt.draw()
    plt.pause(0.01)

def earth_movers_distance(spike_times_a, spike_times_b, rhythm_trace_a, rhythm_trace_b):
    rhythm_total_spike_times_a = len(spike_times_a)
    rhythm_total_spike_times_b = len(spike_times_b)
    cumulative_a = np.cumsum(np.divide(rhythm_trace_a, rhythm_total_spike_times_a))
    cumulative_b = np.cumsum(np.divide(rhythm_trace_b, rhythm_total_spike_times_b))
    # same thing as np.sum(np.abs(np.subtract(cumulative_a, cumulative_b))),
    return  scipy.stats.wasserstein_distance(spike_times_a, spike_times_b)

def victor_purp(onset_times_1, onset_times_2, loop_begin, loop_end):
    q = 1.0 / (10.0 * pq.ms)
    st_a = SpikeTrain(onset_times_1, units='ms', t_start = 0, t_stop= loop_end-loop_begin)
    st_b = SpikeTrain(onset_times_2, units='ms', t_start = 0, t_stop= loop_end-loop_begin)
    vp_f = victor_purpura_distance([st_a, st_b], q)[1][0] # give only the distance between the two different trains (not diagonal)
    return vp_f

# TEST THIS FUNCTION
def compile_unique_interval_list(intervals, tolerance):
    interval_index_lib = np.arange(0, len(intervals))
    unique_interval_list = []
    index = 0
    indices = []
    for interval in intervals: # for every interval
        # if this interval is within tolerance of any item on the list,
        bool_list = [(interval+tolerance > item and interval-tolerance < item) for item in unique_interval_list]
        if any(bool_list):
            indices.append(bool_list.index(True))
            continue # we consider it on the list.
        else: #otherwise
            unique_interval_list.append(interval) # add it to list
            indices.append(len(unique_interval_list) - 1)
    return unique_interval_list, indices

xlsxfiles = []
for file in glob.glob("*.xlsx"):
    if file[0] == '2': # starts with the date, so the type of file we want, i.e., 2021_...
        xlsxfiles.append(file)

xlsxfiles.sort(reverse=True)

### load data ###

worksheet_data_begin_indices = [4, 1]
variables_number = 4   # time_readings, contact trace, stim_onsets, audio_onsets
num_conditions = 3 # pre stim, during stim, post stim

# determine len x axis
 
### open workbook, define worksheets ###
list_of_var_lists = []
list_of_emd_distances = []
list_of_vp_distances = []
list_of_WK_var_lists = []

wb = load_workbook(xlsxfiles[0]) # most recent participant file

for i in range(len(ems_constants.rhythm_strings)):  
    rhythm_substr = ems_constants.rhythm_strings[i]
    
    worksheet = wb[ems_constants.rhythm_strings_names[i]]

    ## how many time readings are there?

    counter = 0
    val = 1.0
    while type(val) is float:
        val = worksheet.cell(row=counter + worksheet_data_begin_indices[0], column=1).value
        counter = counter+1

    time_readings_length = counter-1

    ## make the appropriate size data array and read in values.
    arr = np.empty([time_readings_length, variables_number]) 
    arr[:] = np.NaN

    for r in range(time_readings_length):
        for c in range(variables_number):
            arr[r][c] = worksheet.cell(row=r + worksheet_data_begin_indices[0], column=c + worksheet_data_begin_indices[1]).value

    pre_ems_repeats = worksheet.cell(row=2, column=11).value
    with_ems_repeats = worksheet.cell(row=2, column=12).value
    post_ems_repeats = worksheet.cell(row=2, column=13).value


    contact_x_values = arr[:, 0]
    reading_list = arr[:, 1]
    stim_onsets_temp = arr[:, 2]
    stim_onset_times = stim_onsets_temp[~np.isnan(stim_onsets_temp)] # take care of nans that make array work
    audio_onsets_temp = arr[:, 3]
    audio_onset_times = audio_onsets_temp[~np.isnan(audio_onsets_temp)]

    ## preprocessing
    audio_hold = 30000/ems_constants.bpm 
    x_vec = np.arange(0, np.max(contact_x_values), ems_constants.samp_period_ms)
    stim_trace = spike_times_to_traces(stim_onset_times, ems_constants.actual_stim_length, x_vec, ems_constants.samp_period_ms)
    audio_trace = spike_times_to_traces(audio_onset_times, audio_hold, x_vec, ems_constants.samp_period_ms)

    legend_labels = ["contact trace", "stim trace", "audio trace"]

    # plot_contact_trace_and_rhythm(reading_list, contact_x_values, stim_trace, audio_trace, x_vec, ems_constants.samp_period_ms, legend_labels)

    surpressed_contact_onset_times = process_contact_trace_to_hit_times(reading_list, contact_x_values, ems_constants.baseline_subtractor, ems_constants.surpression_window)

    contact_hold = ems_constants.contact_spike_time_width
    surpressed_contact_trace = spike_times_to_traces(surpressed_contact_onset_times, contact_hold, x_vec, ems_constants.samp_period_ms)

    legend_labels = ["surpressed contact trace", "stim trace", "audio trace"]

    # plot_traces(x_vec, [surpressed_contact_trace, audio_trace, stim_trace], ems_constants.samp_period_ms, legend_labels, ems_constants.rhythm_strings_names[i])
    # plot_contact_trace_and_rhythm(surpressed_contact_trace, x_vec, stim_trace, audio_trace, x_vec, ems_constants.samp_period_ms, legend_labels)

    var_list = [contact_x_values, reading_list, stim_onset_times, audio_onset_times, x_vec, stim_trace, audio_trace, surpressed_contact_onset_times, surpressed_contact_trace]
    list_of_var_lists.append(var_list)

    # analysis
    len_rhythm_ms = len(rhythm_substr) * ems_constants.milliseconds_per_eighthnote
    if ems_constants.metronome_intro_flag == 1:
        len_count_off_ms = len(ems_constants.count_in_substr) * ems_constants.milliseconds_per_eighthnote + 3000
    else:
        len_count_off_ms = 3000
    len_count_off_and_audio_display_ms = len_count_off_ms + pre_ems_repeats * len_rhythm_ms
    len_count_off_and_audio_display_and_ems_ms = len_count_off_ms + pre_ems_repeats *len_rhythm_ms + with_ems_repeats * len_rhythm_ms

    delays_list = [len_count_off_ms, len_count_off_and_audio_display_ms, len_count_off_and_audio_display_and_ems_ms, max(contact_x_values)]
    audio_repeats_distances = []
    ems_repeats_distances = []
    post_ems_repeats_distances = []
    distances_list = []
    vp_dist_list = []
    repeat_list = [pre_ems_repeats, with_ems_repeats, post_ems_repeats]

    # for each loop of rhythm, calculate EMD for contact trace vs real audio rhythm
    # change to spikes
    audio_trace_copy = np.copy(audio_trace)
    for j in range(len(audio_trace)-1):
        if audio_trace_copy[j] == 1:
            audio_trace[j+1] = 0
    
    contact_trace_copy = np.copy(surpressed_contact_trace)
    for j in range(len(surpressed_contact_trace)-1):
        if contact_trace_copy[j] == 1:
            surpressed_contact_trace[j+1] = 0

    for k in range(num_conditions): # for each condition (audio only, ems and audio, post_ems audio only test)
        for j in range(repeat_list[k]): # for each repeat of the rhythm in this condition
            loop_begin = delays_list[k] + j * len_rhythm_ms - 0.5*ems_constants.milliseconds_per_eighthnote #include half an eighthnote before
            loop_end = loop_begin + len_rhythm_ms + 1 * ems_constants.milliseconds_per_eighthnote #include one eighthnote after as well
            contact_bool = np.logical_and((surpressed_contact_onset_times >= loop_begin), (surpressed_contact_onset_times <= loop_end)) # select contact onset times during this loop of rhythm
            audio_bool = np.logical_and((audio_onset_times >= loop_begin), (audio_onset_times <= loop_end)) # select audio onset times during this loop of rhythm
            spike_times_contact = surpressed_contact_onset_times[contact_bool] - loop_begin # how many spikes total?
            spike_times_audio = audio_onset_times[audio_bool] - loop_begin
            trace_selector_bool = np.logical_and((x_vec >= loop_begin), (x_vec <= loop_end)) # which indices in traces are during this loop?
            contact_trace_selected = surpressed_contact_trace[trace_selector_bool] # pick those data points from suprpressed contact trace
            audio_trace_selected = audio_trace[trace_selector_bool] # pick those data points from audio trace
            emd = earth_movers_distance(spike_times_contact, spike_times_audio, contact_trace_selected, audio_trace_selected) # run emd
            distances_list.append(emd) # add to appropriate list.
            vp_dist = victor_purp(spike_times_contact, spike_times_audio, loop_begin, loop_end)
            vp_dist_list.append(vp_dist) 
            title = "total spikes contact: " + str(len(spike_times_contact)) + " total_spikes audio: " + str(len(spike_times_audio)) + " emd = " + str(emd) +  " vic purp: " + str(vp_dist)
            # plot_traces(x_vec[trace_selector_bool], [contact_trace_selected, audio_trace_selected], ems_constants.samp_period_ms, ["contact", "audio"], title)
            
            
    
    list_of_vp_distances.append(vp_dist_list)
    list_of_emd_distances.append(distances_list)
    fig, ax = plt.subplots()
    label_list = ["EMD", "VPD"]
    # normalize EMD
    distances_array = np.array(distances_list)
    distances_array = np.divide(distances_array, np.max(distances_array))
    ax.plot(np.arange(len(distances_array)), distances_array)
    vp_array = np.array(vp_dist_list)
    vp_array = np.divide(vp_array, np.max(vp_array))
    ax.plot(np.arange(len(vp_array)), vp_array)
    ax.legend(label_list)

    ax.set_title("normalized distances for " + rhythm_substr)
    plt.ion()
    plt.show()
    plt.draw()
    plt.pause(0.01)

    # store interval information for wing kris analysis

    # user interval is the time between the nearest contact spike to an audio spike and the previous audio spike.
    # user error is the absolute difference between user interval and the ground truth audio interval.

    # get intervals in this rhythm
    var_lists = []
    for k in range(num_conditions):
        ground_truth_intervals = []
        user_intervals = []
        user_error = []
        this_condition_bools = np.logical_and((audio_onset_times > delays_list[k]), (audio_onset_times < delays_list[k+1]))
        condition_audio_onsets = audio_onset_times[this_condition_bools]
        for j in range(len(condition_audio_onsets)-1):
            gt_interval = condition_audio_onsets[j+1] - condition_audio_onsets[j]
            # get nearest response pulse to audio
            arg_min = np.argmin(np.abs(np.subtract(surpressed_contact_onset_times, condition_audio_onsets[j+1]))) # throws out the first one...
            user_interval = surpressed_contact_onset_times[arg_min] - condition_audio_onsets[j]#response time user - previous audio pulse
            ground_truth_intervals.append(gt_interval)
            user_intervals.append(user_interval)
            user_error.append(np.abs(gt_interval-user_interval))
        var_list = [ground_truth_intervals, user_intervals, user_error]
        var_lists.append(var_list) # now has each WK relevant variable for each condition

    list_of_WK_var_lists.append(var_lists)

print("done")


# organize intervals

slopes = []
intercepts = []
r_sqs = []

# wing kris across rhythms for each condition
for i in range(num_conditions):
    list_of_vars_for_condition = [] # first dim, len_rhythm_types long. second dim, condition, third, ground truth, user, user_error.
    for j in range(len(ems_constants.rhythm_strings)):
        list_of_vars_for_condition.append(list_of_WK_var_lists[j][i])
    ground_truth_intervs_across_rhythms = [el[0] for el in list_of_vars_for_condition] # get all gt intervals
    ground_truth_intervs_across_rhythms_flat = sum(ground_truth_intervs_across_rhythms, [])
    # get list of unique intervals and indices that map all intervals to their assigned unique interval
    unique_gt_intervals, indices = compile_unique_interval_list(ground_truth_intervs_across_rhythms_flat, ems_constants.interval_tolerance)
    user_intervals_across_rhythms = [el[1] for el in list_of_vars_for_condition] # get all user intervasl
    user_intervals_across_rhythms_flat = sum(user_intervals_across_rhythms, [])
    list_of_user_intervals_by_target_interval = [[]] * len(unique_gt_intervals) # make a list of lists as long as unique intervals
    for k in range(len(user_intervals_across_rhythms_flat)): # for every user interval
        target_interval_index = indices[k] # find its unique target 
        # now add it to the list of intervals produced for that target
        #within the list of unique intervals
        list_of_user_intervals_by_target_interval[target_interval_index].append(user_intervals_across_rhythms_flat[k])
    interval_means_list = [sum(item_list)/len(item_list) for item_list in list_of_user_intervals_by_target_interval]
    interval_sd_list = [np.std(item_list) for item_list in list_of_user_intervals_by_target_interval]
    slope, intercept, r_value, p_value, std_err = scipy.stats.linregress(interval_means_list, interval_sd_list) 
    rsq = r_value**2
    slopes.append(slope)
    intercepts.append(intercept)
    r_sqs.append(rsq)

fig, ax = plt.subplots()

legend_labels = ["clock var", "motor var"]
ax.plot(np.arange(len(slopes)), slopes/max(slopes))
ax.plot(np.arange(len(intercepts)), intercepts/max(intercepts))
ax.legend(label_list)

ax.set_title("normalized variances for clock and motor across epochs")
plt.ion()
plt.show()
plt.draw()
plt.pause(0.01)