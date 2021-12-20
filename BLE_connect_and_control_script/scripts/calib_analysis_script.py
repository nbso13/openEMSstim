import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook

LENGTH_BEGIN = 165
LENGTH_END = 270
LENGTH_STEP = 15
INTENSITY_BEGIN = 60
INTENSITY_END = 100
INTENSITY_STEP = 5
BPM = 70
BAUD_RATE = 115200
REPEATS = 1
RHYTH_STR = "101010"
PARTICIPANT_NUMBERS = [0, 1]

# initial results indicate that after finding max level they can take with this max length, max stim length


# notes: we need to think about where we're actuating the muscle along its length - there is a standard for electrode placement. Talk to 
# Lewis about this. This already exists. Jakob might know about this. Probably will.



lengths = range(LENGTH_BEGIN, LENGTH_END, LENGTH_STEP) #length in ms of stimulation pulse -- watch out for mystery numbers - using 130 to 290 as range -- 130 being minimum length of stim and 290 as max necessary for low intensity
# perhaps better to set up a 'settings file'
 # make it global variables
intensities = range(INTENSITY_BEGIN, INTENSITY_END, INTENSITY_STEP) 

rhythm_substr = RHYTH_STR
repeats = REPEATS
bpm = BPM


worksheet_data_begin_indices = [4, 2]
 
### open workbook, define worksheets ###
arr = np.empty((len(intensities), len(lengths)))
arr[:] = np.NaN
data_arr = [arr, np.copy(arr), np.copy(arr)]

wb = load_workbook('2021_11_19_17_48_26_pp2.xlsx')

name_str = ["Pain level", "Actuation level", "Actuation speed"]

for i in range(3):
    worksheet = wb[name_str[i]]
    for r in range(len(intensities)):
        for c in range(len(lengths)):
            data_arr[i][r][c] = worksheet.cell(row=r + worksheet_data_begin_indices[0], column=c + worksheet_data_begin_indices[1]).value



# calibration index

index_val_actuation = np.abs(data_arr[2]-6) # distance from ideal actuation
index_map = np.subtract(data_arr[1],np.add(2*index_val_actuation, data_arr[0]))



### visualize as heat map at the end ###
len_labs = []
for i in lengths:
    len_labs.append(str(i))
int_labs = []
for i in intensities:
    int_labs.append(str(i))

fig, axs = plt.subplots(2, 2)
img = axs[0, 0].imshow(data_arr[0], cmap='hot', interpolation='nearest')
plt.colorbar(img, ax=axs[0,0])
axs[0, 0].set_title('Pain map')
img = axs[0, 1].imshow(data_arr[1], cmap='hot', interpolation='nearest')
plt.colorbar(img, ax=axs[0,1])
axs[0, 1].set_title('Actuation map')
axs[1, 0].imshow(data_arr[2], cmap='hot', interpolation='nearest')
img = plt.colorbar(img, ax=axs[1,0])
axs[1, 0].set_title('Speed map')
img = axs[1,1].imshow(index_map, cmap='hot', interpolation='nearest')
plt.colorbar(img, ax=axs[1,1])
axs[1, 1].set_title('Index Map')

for ax in axs.flat:
    ax.set(xlabel='Stim length (ms)', ylabel='Stim intensity (%)')
    ax.set_xticklabels(len_labs) 
    ax.set_yticklabels(int_labs)
    

plt.show()

wb.close()